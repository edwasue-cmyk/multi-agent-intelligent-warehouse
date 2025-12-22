#!/usr/bin/env python3
# SPDX-FileCopyrightText: Copyright (c) 2025 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: Apache-2.0
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
Comprehensive License and Dependency Audit Tool

This script performs a thorough audit of all dependencies using multiple tools:
- pip-licenses (Python)
- license-checker (Node.js)
- pipdeptree (dependency tree)
- pip-audit (security)
- npm ls (Node.js dependencies)

Generates a comprehensive XLSX report with all findings.
"""

import json
import subprocess
import sys
import os
from pathlib import Path
from typing import Dict, List, Optional, Any
from datetime import datetime
import csv
import tempfile

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False


def run_command(cmd: List[str], cwd: Optional[Path] = None, check: bool = False) -> tuple[bool, str, str]:
    """Run a shell command and return success, stdout, stderr."""
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            cwd=cwd,
            check=check
        )
        return result.returncode == 0, result.stdout, result.stderr
    except Exception as e:
        return False, "", str(e)


def check_tool_installed(tool_name: str, check_cmd: List[str]) -> bool:
    """Check if a tool is installed."""
    success, _, _ = run_command(check_cmd)
    return success


def install_python_tool(tool_name: str) -> bool:
    """Install a Python tool using pip."""
    print(f"Installing {tool_name}...")
    success, stdout, stderr = run_command([sys.executable, "-m", "pip", "install", tool_name])
    if success:
        print(f"✅ {tool_name} installed successfully")
    else:
        print(f"❌ Failed to install {tool_name}: {stderr}")
    return success


def install_npm_tool(tool_name: str, local: bool = True) -> bool:
    """Install an npm tool locally or globally."""
    if local:
        print(f"Installing {tool_name} locally...")
        # Try local install first (no -g flag)
        success, stdout, stderr = run_command(["npm", "install", tool_name])
        if success:
            print(f"✅ {tool_name} installed successfully (local)")
            return True
        else:
            print(f"⚠️ Local install failed, trying global: {stderr}")
    
    print(f"Installing {tool_name} globally (may require sudo)...")
    success, stdout, stderr = run_command(["npm", "install", "-g", tool_name])
    if success:
        print(f"✅ {tool_name} installed successfully (global)")
    else:
        print(f"❌ Failed to install {tool_name}: {stderr}")
    return success


def audit_python_pip_licenses(repo_root: Path) -> List[Dict[str, Any]]:
    """Audit Python packages using pip-licenses."""
    print("\n📦 Auditing Python packages with pip-licenses...")
    
    # Check if tool is installed
    if not check_tool_installed("pip-licenses", ["pip-licenses", "--version"]):
        if not install_python_tool("pip-licenses"):
            return []
    
    # Run pip-licenses
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
        csv_file = f.name
    
    success, stdout, stderr = run_command(
        ["pip-licenses", "--format=csv", f"--output-file={csv_file}"],
        cwd=repo_root
    )
    
    if not success:
        print(f"⚠️ pip-licenses failed: {stderr}")
        return []
    
    # Parse CSV
    packages = []
    try:
        with open(csv_file, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                packages.append({
                    'name': row.get('Name', ''),
                    'version': row.get('Version', ''),
                    'license': row.get('License', ''),
                    'license_text': row.get('LicenseText', ''),
                    'source': 'PyPI',
                    'tool': 'pip-licenses'
                })
    except Exception as e:
        print(f"⚠️ Error parsing pip-licenses CSV: {e}")
    
    # Cleanup
    try:
        os.unlink(csv_file)
    except:
        pass
    
    print(f"✅ Found {len(packages)} Python packages")
    return packages


def audit_python_pipdeptree(repo_root: Path) -> List[Dict[str, Any]]:
    """Get Python dependency tree using pipdeptree."""
    print("\n🌳 Getting Python dependency tree with pipdeptree...")
    
    if not check_tool_installed("pipdeptree", ["pipdeptree", "--version"]):
        if not install_python_tool("pipdeptree"):
            return []
    
    success, stdout, stderr = run_command(["pipdeptree", "--json"], cwd=repo_root)
    
    if not success:
        print(f"⚠️ pipdeptree failed: {stderr}")
        return []
    
    try:
        deptree = json.loads(stdout)
        # Flatten the tree - handle both dict and list formats
        packages = []
        
        if isinstance(deptree, list):
            # List format
            for item in deptree:
                if isinstance(item, dict):
                    pkg_name = item.get('package', {}).get('key', '') if isinstance(item.get('package'), dict) else ''
                    pkg_version = item.get('package', {}).get('installed_version', '') if isinstance(item.get('package'), dict) else ''
                    deps = item.get('dependencies', [])
                    dep_names = [d.get('key', '') if isinstance(d, dict) else str(d) for d in deps]
                    packages.append({
                        'name': pkg_name,
                        'version': pkg_version,
                        'dependencies': ', '.join(dep_names),
                        'tool': 'pipdeptree'
                    })
        elif isinstance(deptree, dict):
            # Dict format
            for pkg_name, pkg_info in deptree.items():
                if isinstance(pkg_info, dict):
                    packages.append({
                        'name': pkg_name,
                        'version': pkg_info.get('installed_version', ''),
                        'dependencies': ', '.join(pkg_info.get('dependencies', {}).keys()) if isinstance(pkg_info.get('dependencies'), dict) else '',
                        'tool': 'pipdeptree'
                    })
        
        print(f"✅ Found {len(packages)} packages in dependency tree")
        return packages
    except Exception as e:
        print(f"⚠️ Error parsing pipdeptree JSON: {e}")
        return []


def audit_python_pip_audit(repo_root: Path) -> List[Dict[str, Any]]:
    """Audit Python packages for security issues using pip-audit."""
    print("\n🔒 Auditing Python packages for security issues with pip-audit...")
    
    if not check_tool_installed("pip-audit", ["pip-audit", "--version"]):
        if not install_python_tool("pip-audit"):
            return []
    
    success, stdout, stderr = run_command(["pip-audit", "--format=json"], cwd=repo_root)
    
    if not success:
        print(f"⚠️ pip-audit failed (may have found vulnerabilities): {stderr}")
        # Still try to parse if there's output
        if not stdout:
            return []
    
    try:
        # pip-audit outputs JSON with dependencies structure
        # Find the JSON object start
        json_start = stdout.find('{')
        if json_start < 0:
            json_start = stdout.find('[')
        
        if json_start >= 0:
            # Try to find the end of the JSON (look for matching braces)
            json_str = stdout[json_start:]
            # Try to parse as much as possible
            try:
                audit_results = json.loads(json_str)
            except json.JSONDecodeError:
                # Try to extract just the vulnerabilities part
                if '"vulns"' in json_str or '"vulnerabilities"' in json_str:
                    # Parse line by line or extract specific sections
                    # For now, try to extract vulnerability info from stderr or parse differently
                    audit_results = {}
                else:
                    raise
        
            vulnerabilities = []
            
            # pip-audit format: {"dependencies": [{"name": "...", "version": "...", "vulns": [...]}]}
            if isinstance(audit_results, dict) and 'dependencies' in audit_results:
                for dep in audit_results.get('dependencies', []):
                    if isinstance(dep, dict):
                        pkg_name = dep.get('name', '')
                        pkg_version = dep.get('version', '')
                        vulns = dep.get('vulns', [])
                        for vuln in vulns:
                            if isinstance(vuln, dict):
                                vulnerabilities.append({
                                    'package': pkg_name,
                                    'installed_version': pkg_version,
                                    'vulnerability_id': vuln.get('id', ''),
                                    'advisory': str(vuln.get('description', '')),
                                    'tool': 'pip-audit'
                                })
            elif isinstance(audit_results, list):
                for vuln in audit_results:
                    if isinstance(vuln, dict):
                        vulnerabilities.append({
                            'package': vuln.get('name', ''),
                            'installed_version': vuln.get('installed_version', ''),
                            'vulnerability_id': vuln.get('vulnerability_id', ''),
                            'advisory': str(vuln.get('advisory', '')),
                            'tool': 'pip-audit'
                        })
            
            print(f"✅ Found {len(vulnerabilities)} security issues")
            return vulnerabilities
        else:
            # No JSON found, but pip-audit may have found issues (check stderr)
            if "vulnerability" in stderr.lower() or "vulnerability" in stdout.lower():
                print("⚠️ pip-audit found vulnerabilities but couldn't parse JSON format")
                print("   Check output manually or use: pip-audit --format=json")
            return []
    except Exception as e:
        print(f"⚠️ Error parsing pip-audit JSON: {e}")
        # Try alternative: run with explicit format
        print("   Attempting alternative format...")
        success2, stdout2, stderr2 = run_command(
            ["pip-audit", "--format=json", "--output=-"],
            cwd=repo_root
        )
        if success2 and stdout2:
            try:
                audit_results = json.loads(stdout2)
                vulnerabilities = []
                if isinstance(audit_results, dict) and 'dependencies' in audit_results:
                    for dep in audit_results.get('dependencies', []):
                        if isinstance(dep, dict):
                            pkg_name = dep.get('name', '')
                            pkg_version = dep.get('version', '')
                            vulns = dep.get('vulns', [])
                            for vuln in vulns:
                                if isinstance(vuln, dict):
                                    vulnerabilities.append({
                                        'package': pkg_name,
                                        'installed_version': pkg_version,
                                        'vulnerability_id': vuln.get('id', ''),
                                        'advisory': str(vuln.get('description', '')),
                                        'tool': 'pip-audit'
                                    })
                print(f"✅ Found {len(vulnerabilities)} security issues (alternative format)")
                return vulnerabilities
            except:
                pass
        return []


def audit_nodejs_license_checker(repo_root: Path) -> List[Dict[str, Any]]:
    """Audit Node.js packages using license-checker."""
    print("\n📦 Auditing Node.js packages with license-checker...")
    
    web_dir = repo_root / "src" / "ui" / "web"
    if not web_dir.exists():
        print("⚠️ Frontend directory not found")
        return []
    
    # Check if tool is installed (try local first, then global)
    local_check = (web_dir / "node_modules" / ".bin" / "license-checker").exists()
    global_check = check_tool_installed("license-checker", ["license-checker", "--version"])
    
    if not local_check and not global_check:
        # Try local install first
        if not install_npm_tool("license-checker", local=True):
            print("⚠️ Could not install license-checker. Skipping Node.js license audit.")
            return []
    
    # Use local version if available, otherwise global
    license_checker_cmd = "license-checker"
    if local_check:
        license_checker_cmd = str(web_dir / "node_modules" / ".bin" / "license-checker")
        # Make sure it's executable
        import stat
        try:
            os.chmod(license_checker_cmd, os.stat(license_checker_cmd).st_mode | stat.S_IEXEC)
        except:
            pass
    
    # Run license-checker
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
        json_file = f.name
    
    # Try using npx if local install didn't work
    if not local_check or not os.path.exists(license_checker_cmd):
        license_checker_cmd = "npx"
        cmd = [license_checker_cmd, "license-checker", "--json", "--out", json_file]
    else:
        cmd = [license_checker_cmd, "--json", "--out", json_file]
    
    success, stdout, stderr = run_command(cmd, cwd=web_dir)
    
    if not success:
        print(f"⚠️ license-checker failed: {stderr[:200]}")
        # Fallback to alternative method
        print("   Trying alternative method...")
        return audit_nodejs_npm_packages_fallback(web_dir)
    
    # Parse JSON
    packages = []
    try:
        with open(json_file, 'r') as f:
            data = json.load(f)
            for pkg_path, pkg_info in data.items():
                # Extract package name from path (e.g., "package@version" -> "package")
                pkg_name = pkg_path.split('@')[0] if '@' in pkg_path else pkg_path
                packages.append({
                    'name': pkg_name,
                    'version': pkg_info.get('version', ''),
                    'license': pkg_info.get('licenses', ''),
                    'license_file': pkg_info.get('licenseFile', ''),
                    'repository': pkg_info.get('repository', ''),
                    'source': 'npm',
                    'tool': 'license-checker'
                })
    except Exception as e:
        print(f"⚠️ Error parsing license-checker JSON: {e}")
    
    # Cleanup
    try:
        os.unlink(json_file)
    except:
        pass
    
    print(f"✅ Found {len(packages)} Node.js packages")
    return packages


def audit_nodejs_npm_packages_fallback(web_dir: Path) -> List[Dict[str, Any]]:
    """Fallback method to get Node.js package info using package.json and npm view."""
    print("   Using npm view as fallback...")
    packages = []
    
    # Read package.json
    package_json = web_dir / "package.json"
    if not package_json.exists():
        return []
    
    try:
        with open(package_json, 'r') as f:
            pkg_data = json.load(f)
        
        all_deps = {}
        all_deps.update(pkg_data.get('dependencies', {}))
        all_deps.update(pkg_data.get('devDependencies', {}))
        
        for pkg_name, version_spec in all_deps.items():
            # Clean version spec (remove ^, ~, etc.)
            version = version_spec.replace('^', '').replace('~', '').replace('>=', '').replace('<=', '')
            
            # Try to get license info from npm
            success, stdout, _ = run_command(
                ["npm", "view", pkg_name, "license", "repository.url", "homepage", "--json"],
                cwd=web_dir
            )
            
            license_info = "Unknown"
            repository = ""
            homepage = ""
            
            if success and stdout:
                try:
                    view_data = json.loads(stdout)
                    if isinstance(view_data, dict):
                        license_info = view_data.get('license', 'Unknown')
                        repo = view_data.get('repository', {})
                        if isinstance(repo, dict):
                            repository = repo.get('url', '')
                        elif isinstance(repo, str):
                            repository = repo
                        homepage = view_data.get('homepage', '')
                    elif isinstance(view_data, str):
                        license_info = view_data
                except:
                    license_info = stdout.strip() if stdout.strip() else "Unknown"
            
            packages.append({
                'name': pkg_name,
                'version': version,
                'license': license_info,
                'repository': repository,
                'homepage': homepage,
                'source': 'npm',
                'tool': 'npm-view (fallback)'
            })
        
        print(f"✅ Found {len(packages)} Node.js packages (fallback method)")
        return packages
    except Exception as e:
        print(f"⚠️ Fallback method failed: {e}")
        return []


def audit_nodejs_npm_ls(repo_root: Path) -> List[Dict[str, Any]]:
    """Get Node.js dependency tree using npm ls."""
    print("\n🌳 Getting Node.js dependency tree with npm ls...")
    
    web_dir = repo_root / "src" / "ui" / "web"
    if not web_dir.exists():
        return []
    
    success, stdout, stderr = run_command(
        ["npm", "ls", "--all", "--json", "--depth=0"],
        cwd=web_dir
    )
    
    if not success:
        print(f"⚠️ npm ls failed: {stderr}")
        return []
    
    try:
        deptree = json.loads(stdout)
        packages = []
        seen = set()  # Track seen packages to avoid infinite recursion
        
        def extract_deps(deps_dict: Dict, depth: int = 0, max_depth: int = 3):
            if not isinstance(deps_dict, dict) or depth > max_depth:
                return
            for name, info in deps_dict.items():
                if not isinstance(info, dict):
                    continue
                # Create unique key
                pkg_key = f"{name}@{info.get('version', '')}"
                if pkg_key in seen:
                    continue
                seen.add(pkg_key)
                
                packages.append({
                    'name': name,
                    'version': info.get('version', ''),
                    'resolved': info.get('resolved', ''),
                    'dependencies': ', '.join(info.get('dependencies', {}).keys()) if isinstance(info.get('dependencies'), dict) else '',
                    'tool': 'npm-ls'
                })
                
                # Recursively process dependencies
                if 'dependencies' in info and isinstance(info['dependencies'], dict):
                    extract_deps(info['dependencies'], depth + 1, max_depth)
        
        if 'dependencies' in deptree and isinstance(deptree['dependencies'], dict):
            extract_deps(deptree['dependencies'])
        
        print(f"✅ Found {len(packages)} packages in npm dependency tree")
        return packages
    except Exception as e:
        print(f"⚠️ Error parsing npm ls JSON: {e}")
        return []


def merge_audit_results(*audit_results: List[List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    """Merge results from multiple audit tools."""
    all_packages = {}
    
    for result_list in audit_results:
        for pkg in result_list:
            name = pkg.get('name', '').lower()
            source = pkg.get('source', 'unknown')
            key = (name, source)
            
            if key not in all_packages:
                all_packages[key] = pkg.copy()
            else:
                # Merge information
                existing = all_packages[key]
                # Update with more complete information
                for k, v in pkg.items():
                    if k not in existing or not existing[k]:
                        existing[k] = v
                    elif k == 'tool' and v not in str(existing.get('tool', '')):
                        existing[k] = f"{existing.get('tool', '')}, {v}"
    
    return list(all_packages.values())


def generate_xlsx_report(audit_results: Dict[str, List[Dict[str, Any]]], output_file: Path):
    """Generate comprehensive XLSX report."""
    if not HAS_OPENPYXL:
        print("❌ openpyxl not installed. Cannot generate XLSX file.")
        print("Install with: pip install openpyxl")
        return False
    
    print(f"\n📊 Generating XLSX report: {output_file}")
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Sheet 1: Python Packages (pip-licenses)
    if 'python_pip_licenses' in audit_results and audit_results['python_pip_licenses']:
        ws = wb.create_sheet("Python Packages")
        headers = ['Package Name', 'Version', 'License', 'License Text', 'Source', 'Tool']
        ws.append(headers)
        
        # Style header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        for pkg in audit_results['python_pip_licenses']:
            ws.append([
                pkg.get('name', ''),
                pkg.get('version', ''),
                pkg.get('license', ''),
                pkg.get('license_text', '')[:500],  # Truncate long text
                pkg.get('source', ''),
                pkg.get('tool', '')
            ])
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # Sheet 2: Node.js Packages (license-checker)
    if 'nodejs_license_checker' in audit_results and audit_results['nodejs_license_checker']:
        ws = wb.create_sheet("Node.js Packages")
        headers = ['Package Name', 'Version', 'License', 'Repository', 'License File', 'Source', 'Tool']
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        for pkg in audit_results['nodejs_license_checker']:
            ws.append([
                pkg.get('name', ''),
                pkg.get('version', ''),
                pkg.get('license', ''),
                pkg.get('repository', ''),
                pkg.get('license_file', ''),
                pkg.get('source', ''),
                pkg.get('tool', '')
            ])
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # Sheet 3: Security Vulnerabilities (pip-audit)
    if 'python_pip_audit' in audit_results and audit_results['python_pip_audit']:
        ws = wb.create_sheet("Security Vulnerabilities")
        headers = ['Package', 'Installed Version', 'Vulnerability ID', 'Advisory', 'Tool']
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
            cell.font = header_font
            cell.alignment = center_align
        
        for vuln in audit_results['python_pip_audit']:
            ws.append([
                vuln.get('package', ''),
                vuln.get('installed_version', ''),
                vuln.get('vulnerability_id', ''),
                vuln.get('advisory', '')[:500],
                vuln.get('tool', '')
            ])
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # Sheet 4: Summary
    ws = wb.create_sheet("Summary")
    ws.append(['Audit Summary', ''])
    ws.append(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws.append(['', ''])
    
    summary_data = [
        ['Category', 'Count'],
        ['Python Packages (pip-licenses)', len(audit_results.get('python_pip_licenses', []))],
        ['Node.js Packages (license-checker)', len(audit_results.get('nodejs_license_checker', []))],
        ['Security Vulnerabilities', len(audit_results.get('python_pip_audit', []))],
        ['Python Dependency Tree Entries', len(audit_results.get('python_pipdeptree', []))],
        ['Node.js Dependency Tree Entries', len(audit_results.get('nodejs_npm_ls', []))],
    ]
    
    for row in summary_data:
        ws.append(row)
    
    # Style summary header
    for col_num in range(1, 3):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save workbook
    wb.save(output_file)
    print(f"✅ XLSX report generated: {output_file}")
    return True


def main():
    """Main audit function."""
    repo_root = Path(__file__).parent.parent.parent
    
    print("=" * 70)
    print("Comprehensive License and Dependency Audit")
    print("=" * 70)
    print(f"Repository: {repo_root}")
    print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)
    
    # Run all audits
    audit_results = {
        'python_pip_licenses': audit_python_pip_licenses(repo_root),
        'python_pipdeptree': audit_python_pipdeptree(repo_root),
        'python_pip_audit': audit_python_pip_audit(repo_root),
        'nodejs_license_checker': audit_nodejs_license_checker(repo_root),
        'nodejs_npm_ls': audit_nodejs_npm_ls(repo_root),
    }
    
    # Generate XLSX report
    output_file = repo_root / "docs" / "License_Audit_Report.xlsx"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    success = generate_xlsx_report(audit_results, output_file)
    
    if success:
        print("\n" + "=" * 70)
        print("✅ Audit Complete!")
        print(f"📊 Report saved to: {output_file}")
        print("=" * 70)
    else:
        print("\n❌ Failed to generate XLSX report")
        sys.exit(1)


if __name__ == "__main__":
    main()

