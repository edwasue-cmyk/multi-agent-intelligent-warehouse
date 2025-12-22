#!/usr/bin/env python3
# SPDX-FileCopyrightText: Copyright (c) 2025 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: Apache-2.0
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
Generate Excel file from SOFTWARE_INVENTORY.md
Extracts all package information and creates an Excel spreadsheet.
"""

import re
import sys
from pathlib import Path
from typing import List, Dict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def parse_markdown_table(content: str, section_start: str, section_end: str) -> List[Dict[str, str]]:
    """Parse markdown table from content between section markers."""
    # Find the section
    start_idx = content.find(section_start)
    if start_idx == -1:
        return []
    
    end_idx = content.find(section_end, start_idx)
    if end_idx == -1:
        section_content = content[start_idx:]
    else:
        section_content = content[start_idx:end_idx]
    
    # Find the table
    table_start = section_content.find('| Package Name')
    if table_start == -1:
        return []
    
    # Extract table lines
    lines = section_content[table_start:].split('\n')
    packages = []
    
    # Skip header and separator lines
    for line in lines[2:]:  # Skip header and separator
        line = line.strip()
        if not line or not line.startswith('|'):
            continue
        
        # Parse table row
        parts = [p.strip() for p in line.split('|')[1:-1]]  # Remove empty first/last
        if len(parts) >= 8:
            packages.append({
                'Package Name': parts[0],
                'Version': parts[1],
                'License': parts[2],
                'License URL': parts[3],
                'Author': parts[4],
                'Source': parts[5],  # Location where component was downloaded
                'Distribution Method': parts[6],
                'Download Location': parts[7]
            })
    
    return packages


def create_excel_file(packages: List[Dict[str, str]], output_file: Path):
    """Create Excel file with package inventory."""
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Package Inventory"
    
    # Define headers (matching user's requested columns)
    headers = [
        'Package Name',
        'Version',
        'License',
        'License URL',
        'Author',
        'Location where component was downloaded',  # Maps to 'Source'
        'Distribution Method'
    ]
    
    # Write headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Write package data
    for row_idx, pkg in enumerate(packages, start=2):
        ws.cell(row=row_idx, column=1, value=pkg.get('Package Name', ''))
        ws.cell(row=row_idx, column=2, value=pkg.get('Version', ''))
        ws.cell(row=row_idx, column=3, value=pkg.get('License', ''))
        ws.cell(row=row_idx, column=4, value=pkg.get('License URL', ''))
        ws.cell(row=row_idx, column=5, value=pkg.get('Author', ''))
        # Map 'Source' to 'Location where component was downloaded'
        ws.cell(row=row_idx, column=6, value=pkg.get('Source', ''))
        ws.cell(row=row_idx, column=7, value=pkg.get('Distribution Method', ''))
    
    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)  # Cap at 100 characters
        ws.column_dimensions[column].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws['A1'] = 'Package Inventory Summary'
    summary_ws['A1'].font = Font(bold=True, size=14)
    
    summary_ws['A3'] = 'Total Packages:'
    summary_ws['B3'] = len(packages)
    summary_ws['A3'].font = Font(bold=True)
    
    summary_ws['A4'] = 'Python Packages:'
    python_count = sum(1 for p in packages if p.get('Distribution Method', '').lower() == 'pip')
    summary_ws['B4'] = python_count
    summary_ws['A4'].font = Font(bold=True)
    
    summary_ws['A5'] = 'Node.js Packages:'
    node_count = sum(1 for p in packages if p.get('Distribution Method', '').lower() == 'npm')
    summary_ws['B5'] = node_count
    summary_ws['A5'].font = Font(bold=True)
    
    # Save workbook
    wb.save(output_file)
    print(f"✅ Excel file created: {output_file}")
    print(f"   Total packages: {len(packages)}")
    print(f"   Python packages: {python_count}")
    print(f"   Node.js packages: {node_count}")


def main():
    """Generate Excel file from SOFTWARE_INVENTORY.md."""
    repo_root = Path(__file__).parent.parent.parent
    inventory_file = repo_root / 'docs' / 'SOFTWARE_INVENTORY.md'
    output_file = repo_root / 'docs' / 'Package_Inventory.xlsx'
    
    if not inventory_file.exists():
        print(f"❌ Error: {inventory_file} not found")
        sys.exit(1)
    
    # Read inventory file
    with open(inventory_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Parse Python packages
    python_packages = parse_markdown_table(
        content,
        '## Python Packages (PyPI)',
        '## Node.js Packages'
    )
    
    # Parse Node.js packages
    node_packages = parse_markdown_table(
        content,
        '## Node.js Packages (npm)',
        '## Notes'
    )
    
    # Combine all packages
    all_packages = python_packages + node_packages
    
    if not all_packages:
        print("❌ Error: No packages found in SOFTWARE_INVENTORY.md")
        sys.exit(1)
    
    # Create Excel file
    create_excel_file(all_packages, output_file)
    
    print(f"\n📊 Package breakdown:")
    print(f"   Python (PyPI): {len(python_packages)}")
    print(f"   Node.js (npm): {len(node_packages)}")
    print(f"   Total: {len(all_packages)}")


if __name__ == '__main__':
    main()

